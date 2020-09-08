#-*- coding: utf-8 -*-
#@Time    :2020/9/712:14
#@Author  :caofade
#@Email   :m18315971975@163.com
#@File    :python6.py
#@QQ      :1139903891
#@Software:PyCharm
#接口自动化步骤
#1接口测试用例
#2python代码读取接口测试用例
#3requests库发送接口请求
#4执行结果与预期结果进行比对，用例是否通过
#5结果回写到excel里
'''
代码自动读取测试数据+自动回写数据===测试用例一般excel居多
第三方库：openpyxl--读取回写
1.安装：pip install oenpyxl
2.导入
excel表格操作：三大对象
1.工作簿对象，直接把文件托到pycharm里的文件单里去
2.表单---sheet
3.单元格---cell
#定义函数的三步骤，1.实现功能2.参数变化的值设置变量3.返回值
'''
import openpyxl       #导入第三方库
import requests
#读取数据
def read_data(filename,sheetname):
    wb=openpyxl.load_workbook(filename)   #加载工作簿对象，并且赋值变量
    sheet=wb[sheetname]   #获取表单
    case_list=[]     #装用例的大列表
    max_row=sheet.max_row    #获取最大行数
    for i in range(2,max_row+1):
        case=dict(
        case_id=sheet.cell(row=i,column=1).value,
        url=sheet.cell(row=i,column=5).value,      #行和列找到对应单元格
        #print(url.value)      #通过单元格.value---里面的内容
        data=sheet.cell(row=i,column=6).value,    #参数
        expected=sheet.cell(row=i,column=7).value    #期望结果
        )    #大字典内每个小字典都是一条测试用例
        case_list.append(case)
    return case_list
cases=read_data("test_case_api.xlsx","login")
print(cases)
#发送接口请求
def api_request(api_url,api_data):
    qcd_headers_register={"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    response=requests.post(url=api_url,json=api_data,headers=qcd_headers_register)
    return response.json()



#回写
def write_result(filename,sheetname,row,column,final_result):
    wb=openpyxl.load_workbook(filename)   #加载工作簿对象，并且赋值变量
    sheet=wb[sheetname]
    sheet.cell(row=row,column=column).value = final_result
    wb.save(filename)       #保存，写入数据才会生效

def execute_func(filename,sheetname):
    cases=read_data(filename,sheetname)
    for case in cases:
        case_id = case.get("case_id")
        url = case.get("url")
        data = case["data"]
        data = eval(data)       #脱引号
        expected = case.get("expected")
        real_result = api_request(api_url=url,api_data=data)     #执行结果
    #    print(real_result)
    #对比
        real_msg = real_result["msg"]
        expected = eval(expected)
        expected_msg = expected["msg"]
        print("真实执行结果是：{}".format(real_msg))
        if real_msg == expected_msg:
            print("第{}条测试用例通过！".format(case_id))
            final_result ="passed"
        else:
            print("第{}条测试用例不通过！".format(case_id))
            final_result = "failed"
        print("*"*20)
        write_result(filename, sheetname, case_id+1, 8, final_result)
execute_func("test_case_api.xlsx","register")